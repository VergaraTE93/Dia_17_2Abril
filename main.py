import openpyxl
#creacion del archivo
archivo1= openpyxl.Workbook()
hojacalc1= archivo1.active
print(f'hoja activa: {hojacalc1.title}')
hojacalc1.title= "venta producto"
print(f'hoja activa: {archivo1.active.title}')
#modifica el nombre de la hoja de calcula
# puedo crear otra hoja con create_sheet
hoja2= archivo1.create_sheet("hoja2")
print(archivo1.sheetnames)
archivo1= openpyxl.Workbook()
hojacalc1= archivo1.active
hojacalc1["A1"]= "Mes"
hojacalc1["B1"]= "Cantidad produto"
hojacalc1["C1"]= "Valor"
hojacalc1["A2"]= "Enero"
hojacalc1["A3"]= "Febrero"
hojacalc1["A4"]= "Marzo"
hojacalc1['B2']= 15
hojacalc1['B3']= 50
hojacalc1['B4']= 30
hojacalc1['C2']= 100500
hojacalc1['C3']= 335000
hojacalc1['C4']= 201000
#creacion de el contrenido de la oja de excel dentro del archivo
archivo1.save("Ventas_de_cuido_primer _trimestre_excel.xlsx")
#con save guardamos lo que se introdujo en la oja de excel
archivo1= openpyxl.load_workbook('Ventas_de_cuido_primer _trimestre_excel.xlsx')
#con el anterior load puedes abrir el archivo que acabamos de crear
# ahora imprimimos para leerlo
print(hojacalc1["A1"].value)
print(hojacalc1["A2"].value)
print(hojacalc1["A3"].value)
print(hojacalc1["A4"].value)
print(hojacalc1["B1"].value)
print(hojacalc1["B2"].value)
print(hojacalc1["B3"].value)
print(hojacalc1["B4"].value)
print(hojacalc1["C1"].value)
print(hojacalc1["C2"].value)
print(hojacalc1["C3"].value)
print(hojacalc1["C4"].value)
